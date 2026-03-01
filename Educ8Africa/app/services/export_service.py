# app/services/export_service.py
import io
import csv
from openpyxl import Workbook
from sqlalchemy.orm import Session
from app.models.user import User, UserRole
from app.models.payment import Payment
from app.models.registration import Registration
from app.models.course import Course
from app.models.referral import Referral
from fastapi.responses import StreamingResponse


class ExportService:
    """Handles data export operations for admin."""

    @staticmethod
    def export_students(db: Session, format: str = "csv") -> StreamingResponse:
        """Export all registered students."""
        students = (
            db.query(User)
            .filter(User.role == UserRole.USER)
            .order_by(User.created_at.desc())
            .all()
        )

        headers = [
            "ID",
            "Full Name",
            "Email",
            "Phone Number",
            "Referral Code",
            "Status",
            "Registered At",
        ]
        rows = []
        for s in students:
            rows.append(
                [
                    str(s.id),
                    s.full_name,
                    s.email,
                    s.phone_number,
                    s.referral_code,
                    "Active" if s.is_active else "Inactive",
                    str(s.created_at),
                ]
            )

        if format == "excel":
            return ExportService._generate_excel(
                "students", headers, rows
            )
        return ExportService._generate_csv("students", headers, rows)

    @staticmethod
    def export_payments(db: Session, format: str = "csv") -> StreamingResponse:
        """Export all payment records."""
        payments = (
            db.query(Payment)
            .order_by(Payment.created_at.desc())
            .all()
        )

        headers = [
            "ID",
            "Student Name",
            "Student Email",
            "Course",
            "Amount",
            "Reference",
            "Status",
            "Paid At",
            "Created At",
        ]
        rows = []
        for p in payments:
            user = db.query(User).filter(User.id == p.user_id).first()
            course = db.query(Course).filter(Course.id == p.course_id).first()
            rows.append(
                [
                    str(p.id),
                    user.full_name if user else "",
                    user.email if user else "",
                    course.course_name if course else "",
                    str(p.amount),
                    p.reference,
                    p.status.value if hasattr(p.status, "value") else str(p.status),
                    str(p.paid_at) if p.paid_at else "",
                    str(p.created_at),
                ]
            )

        if format == "excel":
            return ExportService._generate_excel(
                "payments", headers, rows
            )
        return ExportService._generate_csv("payments", headers, rows)

    @staticmethod
    def export_referrals(
        db: Session, format: str = "csv"
    ) -> StreamingResponse:
        """Export all referral records."""
        referrals = db.query(Referral).all()

        headers = [
            "ID",
            "Referrer Name",
            "Referrer Email",
            "Referred User",
            "Referred Email",
            "Course",
            "Commission",
            "Paid",
            "Created At",
        ]
        rows = []
        for r in referrals:
            referrer = (
                db.query(User).filter(User.id == r.referrer_id).first()
            )
            referred = (
                db.query(User).filter(User.id == r.referred_user_id).first()
            )
            course = (
                db.query(Course).filter(Course.id == r.course_id).first()
            )
            rows.append(
                [
                    str(r.id),
                    referrer.full_name if referrer else "",
                    referrer.email if referrer else "",
                    referred.full_name if referred else "",
                    referred.email if referred else "",
                    course.course_name if course else "",
                    str(r.commission),
                    "Yes" if r.is_paid else "No",
                    str(r.created_at),
                ]
            )

        if format == "excel":
            return ExportService._generate_excel(
                "referrals", headers, rows
            )
        return ExportService._generate_csv("referrals", headers, rows)

    @staticmethod
    def _generate_csv(
        filename: str, headers: list, rows: list
    ) -> StreamingResponse:
        """Generate a CSV file as a streaming response."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.csv"
            },
        )

    @staticmethod
    def _generate_excel(
        filename: str, headers: list, rows: list
    ) -> StreamingResponse:
        """Generate an Excel file as a streaming response."""
        wb = Workbook()
        ws = wb.active
        ws.title = filename.capitalize()

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = cell.font.copy(bold=True)

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}.xlsx"
            },
        )